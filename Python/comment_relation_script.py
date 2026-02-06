import openpyxl
from collections import defaultdict

from datetime import datetime

###
# JOURNALS - Data manipulation ->

# Filter Notes for BLANKS and DELETE those rows, if kept they will enter User ID of Commentor in comments 

# You can also format files prior for [Custom] -> dd/MMM/yy h:mm AM/PM

# Add column to the right of "created at" -> Remove column A & B 

# [journable_id    user_id    notes    created_at    Column1]
# In first cell =TEXTJOIN(";",TRUE,TEXT(D2,"dd/MMM/yy h:mm AM/PM"),B2,C2)

# Do a fill down for the whole column

# Make another blank column and for the TEXTJOIN column and COPY the formula results as values to the new column





now = datetime.now()
formatted_date = now.strftime("%Y-%m-%d-%H-%M")
print(f"Formatted Date: {formatted_date}")
# Load workbook and worksheets
wb = openpyxl.load_workbook(r"C:\projects\sbrown\Python\db_file_exports\output_csv_with_additional_columns\second_updated_workbook-relations.xlsx")
master_ws = wb["work_packages_202602052105"]
control_ws = wb["journals_202602052106"]

# Step 1: Map Master column headers (Row 1) to all matching columns
header_map = defaultdict(list)
for col in range(1, master_ws.max_column + 1):
    val = master_ws.cell(row=1, column=col).value
    if val:
        header_map[str(val).strip()].append(col)

# Step 2: Map Master row keys (Col A) to row numbers
row_map = {}
for row in range(2, master_ws.max_row + 1):
    key = master_ws.cell(row=row, column=1).value
    if key:
        row_map[str(key).strip()] = row

# Step 3: Track usage of (row_key, col_header) pair
usage_tracker = defaultdict(int)

# Step 4: Process each row in Control sheet safely
for row in range(2, control_ws.max_row + 1):
    row_key = control_ws.cell(row=row, column=1).value       # Column A
    value_to_insert = control_ws.cell(row=row, column=6).value  # Column E
    col_header = "comment"   # Column C

    # Skip if row is blank or incomplete
    if not row_key or not col_header or value_to_insert is None:
        continue

    row_key_str = str(row_key).strip()
    col_header_str = str(col_header).strip()

    # Find target row and column(s)
    target_row = row_map.get(row_key_str)
    matching_cols = header_map.get(col_header_str)

    if target_row and matching_cols:
        usage_count = usage_tracker[(row_key_str, col_header_str)]
        if usage_count < len(matching_cols):
            target_col = matching_cols[usage_count]
            master_ws.cell(row=target_row, column=target_col).value = value_to_insert
            usage_tracker[(row_key_str, col_header_str)] += 1
        else:
            print(f"⚠️ All '{col_header_str}' columns used for row_key '{row_key_str}'")
    else:
        if not target_row:
            print(f"⚠️ Row key '{row_key_str}' not found in Master sheet")
        if not matching_cols:
            print(f"⚠️ Column header '{col_header_str}' not found in Master sheet")

# Save workbook
wb.save(r"C:\projects\sbrown\Python\db_file_exports\output_csv_with_additional_columns\second_updated_workbook-relations-comments.xlsx")
print("✅ Done. Saved as 'second_updated_workbook-relations-comments.xlsx'")
now = datetime.now()
formatted_date = now.strftime("%Y-%m-%d-%H-%M")
print(f"Formatted Date: {formatted_date}")

import openpyxl
from collections import defaultdict
from datetime import datetime


now = datetime.now()
formatted_date = now.strftime("%Y-%m-%d-%H-%M")
print(f"Formatted Date: {formatted_date}")
# Load workbook and worksheets
wb = openpyxl.load_workbook("C:\\projects\\sbrown\\Python\\temp_exports_for_db\\fullworkbook_second_attempt.xlsx")
master_ws = wb["Sheet1"]
control_ws = wb["relations_202601280944_non_asci"]


# Step 1: Map headers in Master to all column indexes where they appear
header_row = 1
header_map = defaultdict(list)

for col in range(1, master_ws.max_column + 1):
    header_value = master_ws.cell(row=header_row, column=col).value
    if header_value:
        header_map[str(header_value).strip()].append(col)

# Step 2: Map row keys in Master sheet (Column A) to row numbers
row_map = {}
for row in range(2, master_ws.max_row + 1):
    row_key = master_ws.cell(row=row, column=1).value
    if row_key:
        row_map[str(row_key).strip()] = row

# Step 3: Track how many times each (row_key, col_header) pair has been used
usage_counter = defaultdict(int)

# Step 4: Iterate over Control sheet rows
row = 2
while True:
    row_key = control_ws.cell(row=row, column=2).value
    value_to_insert = control_ws.cell(row=row, column=3).value
    col_header = control_ws.cell(row=row, column=7).value

    if row_key is None and value_to_insert is None and col_header is None:
        break  # End of meaningful data

    if row_key and value_to_insert is not None and col_header:
        row_key_str = str(row_key).strip()
        col_header_str = str(col_header).strip()

        # Get row in master sheet
        target_row = row_map.get(row_key_str)
        # Get list of matching column indexes
        matching_cols = header_map.get(col_header_str)

        if target_row and matching_cols:
            # Get how many times this (row_key, col_header) has been used
            count = usage_counter[(row_key_str, col_header_str)]

            if count < len(matching_cols):
                target_col = matching_cols[count]
                master_ws.cell(row=target_row, column=target_col).value = value_to_insert
                usage_counter[(row_key_str, col_header_str)] += 1
            else:
                print(f"⚠️ No more available '{col_header_str}' columns for row_key '{row_key_str}'")
        else:
            print(f"⚠️ Could not find row '{row_key_str}' or column '{col_header_str}' in Master")

    row += 1

# Save workbook
wb.save("C:\\projects\\sbrown\\Python\\temp_exports_for_db\\second_updated_workbook-relations.xlsx")
print("✅ Master sheet updated and saved as 'second_updated_workbook.xlsx'")
now = datetime.now()
formatted_date = now.strftime("%Y-%m-%d-%H-%M")
print(f"Formatted Date: {formatted_date}")